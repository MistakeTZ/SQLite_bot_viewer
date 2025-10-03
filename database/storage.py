import csv
from datetime import datetime, timedelta
import io
import os
from sqlite3 import Connection, connect
from typing import Dict
import logging

from tabulate import tabulate
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from tasks.config import tz
from tasks.loader import sender


class Database:
    db: Connection
    table_data = {}
    last_query = None

    def __init__(self, mem_db, name):
        """Create a new Database object."""
        self.db = mem_db
        self.name = name.split(".")[0]

        self.tables = [
            table[0] for table in self.db.execute(
                "SELECT name FROM sqlite_master WHERE type='table'",
            ).fetchall()
        ]

        self.update_time = datetime.now(tz=tz)

    def unload(self):
        self.db.close()

    def get_query(self, query):
        self.update_time = datetime.now(tz=tz)
        execution = self.db.execute(query)
        description = [row[0] for row in execution.description]

        self.last_query = query

        return [list(row) for row in execution.fetchall()], description

    def execute_query(self, query):
        self.update_time = datetime.now(tz=tz)
        self.db.execute(query)
        self.db.commit()

    def get_table(self, table):
        table_data = self.table_data.get(table)
        if not table_data:
            table_data = self.get_query(f"SELECT * FROM {table}")
            self.table_data[table] = table_data

        return self.tabulate_result(*table_data)

    def tabulate_result(self, tab_list, headers):
        width, height = len(tab_list), len(headers)
        for row in range(width):
            for cell in range(height):
                if tab_list[row][cell] is None:
                    tab_list[row][cell] = ""

        return tabulate(
            tab_list,
            headers=headers,
            tablefmt="grid",
            maxcolwidths=14,
        )

    def get_buttons(self):
        buttons = [
            [table, f"table_{i}"] for i, table in enumerate(self.tables)
        ]
        buttons += [
            [sender.text("get_sqlite"), "get_sqlite"],
            [sender.text("get_excel"), "get_excel"],
            [sender.text("get_csv"), "get_csv"],
            [sender.text("all_tables"), "back"],
        ]

        return buttons

    def get_sqlite(self):
        """Return a binary copy of the SQLite database."""
        buffer = io.BytesIO()
        with connect(":memory:") as mem_db:
            self.db.backup(mem_db)

            if hasattr(mem_db, "serialize"):
                buffer.write(mem_db.serialize())
            else:
                mem_db.backup(
                    connect("file:memdb1?mode=memory&cache=shared", uri=True),
                )
                for line in mem_db.iterdump():
                    buffer.write(f"{line}\n".encode())

        buffer.seek(0)

        return buffer.getvalue()

    def get_csv(self):
        file_path = f"{self.name}.csv"

        if not self.last_query:
            raise Exception("No last query")

        values, header = self.get_query(self.last_query)

        with open(os.path.join("temp", file_path), "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(values)

        return file_path

    def get_excel(self):
        # Формируем имя файла
        file_path = f"{self.name}.xlsx"

        if self.last_query:
            tables = [""]
            query = self.last_query
        else:
            tables = self.tables
            query = "SELECT * FROM "

        # Создаем Excel-файл
        with pd.ExcelWriter(
            os.path.join("temp", file_path),
            engine="openpyxl",
        ) as writer:
            for table in tables:
                try:
                    # Получаем данные таблицы
                    df = pd.read_sql_query(query + table, self.db)

                    # Записываем данные в лист
                    sheet_name = table[:31] or "Sheet1"
                    df.to_excel(writer, index=False, sheet_name=sheet_name)

                    # Получаем объекты для форматирования
                    worksheet = writer.sheets[sheet_name]

                    # Стили оформления
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(
                        start_color="4F81BD",
                        end_color="4F81BD",
                        fill_type="solid",
                    )
                    cell_alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True)
                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # Форматируем заголовки
                    for col_num, column_name in enumerate(df.columns, 1):
                        col_letter = get_column_letter(col_num)

                        # Заголовок
                        header_cell = worksheet[f"{col_letter}1"]
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                        header_cell.alignment = Alignment(
                            horizontal="center", vertical="center")
                        header_cell.border = thin_border

                        # Автоподбор ширины столбца
                        max_length = max(
                            df[column_name].astype(str).str.len().max(),
                            len(str(column_name)),
                        )
                        worksheet.column_dimensions[col_letter].width = min(
                            max_length + 2, 30)

                    # Форматируем ячейки с данными
                    for row in worksheet.iter_rows(
                        min_row=2,
                        max_row=len(df) + 1,
                    ):
                        for cell in row:
                            cell.alignment = cell_alignment
                            cell.border = thin_border

                    # Замораживаем заголовки и добавляем фильтры
                    worksheet.freeze_panes = "A2"
                    worksheet.auto_filter.ref = worksheet.dimensions

                except Exception as e:
                    logging.error(
                        f"Ошибка при обработке таблицы {table}: {str(e)}",
                    )
                    continue

        return file_path

    def __str__(self):
        return sender.text("db_data", self.name, ", ".join(self.tables))


def clear_databases(now):
    for user_id, db in databases.items():
        if now - db.update_time > timedelta(minutes=30):
            db.unload()
            del databases[user_id]


databases: Dict[int, Database] = {}
